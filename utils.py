import cv2 # OpenCV
import pyautogui # Automatização de tarefas
import numpy as np # Manipulação de arrays
from datetime import datetime # Manipulação de datas e horas
import time # Manipulação de tempo
import customtkinter as ctk # Módulo para criar interfaces gráficas
import json # Manipulação de arquivos JSON
import os # Manipulação de arquivos e pastas
import xlrd # Leitura de arquivos Excel
from xlutils.copy import copy # Cópia de arquivos Excel
from openpyxl import Workbook # Manipulação de arquivos Excel
import sys # Manipulação do sistema
import calendar # Manipulação de calendários
import logging # Registro de logs

class Utils: # Classe com funções úteis

    def show_message(title, message): # Função para exibir uma mensagem
        try:
            window = ctk.CTkToplevel()
            window.title(title)
            window.geometry("300x150") # Define o tamanho da janela

            label = ctk.CTkLabel(window, text=message, wraplength=250) # Cria um rótulo com a mensagem
            label.pack(pady=20) # Adiciona o rótulo à janela

            button = ctk.CTkButton(window, text="OK", command=window.destroy) # Cria um botão para fechar a janela
            button.pack(pady=10)

            window.grab_set()  # Torna a janela modal
        except Exception as e:
            logging.error(f"Ocorreu um erro ao exibir a mensagem CtkTopLevel: {e}")
            raise


    def clicar_elemento(imagem, descricao, imagem_alternativa=None, tentativas=3, confidence=0.9): # Função para clicar em um elemento na tela
        def localizar_imagem_na_tela(imagem_caminho, confidence):
            # Captura a tela
            screenshot = pyautogui.screenshot()
            # Converte a captura de tela para um formato que o OpenCV pode usar
            screenshot = np.array(screenshot)
            screenshot = cv2.cvtColor(screenshot, cv2.COLOR_RGB2BGR)
            
            # Carrega a imagem que será buscada
            template = cv2.imread(imagem_caminho, cv2.IMREAD_COLOR)
            
            if template is None:
                logging.error(f"Não foi possível carregar a imagem {imagem_caminho}")
                raise

            # Realiza a correspondência entre a imagem e a tela
            result = cv2.matchTemplate(screenshot, template, cv2.TM_CCOEFF_NORMED)
            min_val, max_val, min_loc, max_loc = cv2.minMaxLoc(result)

            # Verifica se a correspondência é maior ou igual ao confidence desejado
            if max_val >= confidence:
                template_height, template_width = template.shape[:2]
                center_x = max_loc[0] + template_width // 2
                center_y = max_loc[1] + template_height // 2
                return (center_x, center_y)
            else:
                raise

        # Tentativa de encontrar e clicar na imagem principal
        for tentativa in range(tentativas):
            try:
                location = localizar_imagem_na_tela(imagem, confidence)
                if location:
                    pyautogui.click(location)
                    logging.info(f"{descricao} encontrado(a) e clicado(a) com sucesso.")
                    return
                else:
                    logging.info(f"{descricao} não encontrado(a) na tela. Tentativa {tentativa + 1} de {tentativas}")

            except Exception as e:
                logging.error(f"Ocorreu um erro ao clicar em {descricao}: {e}")
                logging.info(f"{descricao} não encontrado(a) na tela. Tentativa {tentativa + 1} de {tentativas}")

            time.sleep(20)

        # Se a imagem principal não for encontrada, tenta a imagem alternativa (se fornecida)
        if imagem_alternativa: 
            logging.info(f"Tentando imagem alternativa para {descricao}.")
            for tentativa in range(tentativas):
                try:
                    location = localizar_imagem_na_tela(imagem_alternativa, confidence)
                    if location:
                        pyautogui.click(location)
                        logging.info(f"Imagem alternativa para {descricao} encontrada e clicada com sucesso.")
                        return
                    else:
                        logging.info(f"Imagem alternativa para {descricao} não encontrada na tela. Tentativa {tentativa + 1} de {tentativas}")

                except Exception as e:
                    logging.error(f"Ocorreu um erro ao clicar na imagem alternativa para {descricao}: {e}")
                    logging.info(f"Imagem alternativa para {descricao} não encontrada na tela. Tentativa {tentativa + 1} de {tentativas}")

                time.sleep(15)

        # Se todas as tentativas falharem
        logging.error(f"{descricao} e a imagem alternativa não foram encontrados após {tentativas} tentativas.")
        raise Exception(f"{descricao} e a imagem alternativa não foram encontrados.")

    @staticmethod
    def salvar_caminho_relatorios(entries): # Função para salvar o caminho dos relatórios
        software_path = entries['software_path_entry'].get() # Obtém o caminho do software
        arguments = entries['arguments_entry'].get()
        name_sheets = entries['name_sheets_entry'].get()
        caminho_mrp_desktop = entries['caminho_mrp_desktop_entry'].get()
        caminho = entries['caminho_entry'].get()
        caminho_parametros = entries['caminho_parametros_entry'].get()
        caminho_mov_odf = entries['caminho_mov_odf_entry'].get()
        caminho_faturamento_peso = entries['caminho_faturamento_peso_entry'].get()
        caminho_mrp = entries['caminho_mrp_entry'].get()

        # Cria um dicionário com as configurações
        config = { 
            "software_path": software_path,
            "arguments": arguments,
            "name_sheets": name_sheets,
            "caminho_mrp_desktop": caminho_mrp_desktop,
            "caminho": caminho,
            "caminho_parametros": caminho_parametros,
            "caminho_mov_odf": caminho_mov_odf,
            "caminho_faturamento_peso": caminho_faturamento_peso,
            "caminho_mrp": caminho_mrp
        }

        # Salvar o caminho e os argumentos em um arquivo JSON
        try: 
            with open('configuracoes_software.json', 'w') as f: # Abre o arquivo de configurações
                json.dump(config, f, indent=4)
            Utils.show_message("Sucesso", "Caminho do software, argumentos e relatorios salvos com sucesso.")
        except Exception as e:
            Utils.show_message("Erro", f"Erro ao salvar as configurações de sofware: {str(e)}")
            logging.error(f"Erro ao salvar as configurações de software: {str(e)}")
            raise

    @staticmethod
    def salvar_nome_arquivos(entries): # Função para salvar o nome dos arquivos
        estoque_supper = entries['estoque_supper_entry'].get()
        faturamento_peso = entries['faturamento_peso_entry'].get()
        pedido_compra = entries['pedido_compra_entry'].get()
        req_almoxarifado = entries['req_almoxarifado_entry'].get()
        mov_saida_odf = entries['mov_saida_odf_entry'].get()
        mrp_filial_dois = entries['mrp_filial_dois_entry'].get()
        mrp_jms = entries['mrp_jms_entry'].get()
        mrp_jm = entries['mrp_jm_entry'].get()
        parametros = entries['parametros_entry'].get()

        # Cria um dicionário com as configurações
        config = {
            "estoque_supper": estoque_supper, # Salva o nome do arquivo de estoque supper
            "faturamento_peso": faturamento_peso,
            "pedido_compra": pedido_compra,
            "req_almoxarifado": req_almoxarifado,
            "mov_saida_odf": mov_saida_odf,
            "mrp_filial_dois": mrp_filial_dois,
            "mrp_jms": mrp_jms,
            "mrp_jm": mrp_jm,
            "parametros": parametros
        }

        # Salvar o caminho e os argumentos em um arquivo JSON
        try:
            with open('configuracoes_nomes.json', 'w') as f:
                json.dump(config, f, indent=4)
            Utils.show_message("Sucesso", "o nome dos arquivos foram salvos.")
        except Exception as e:
            Utils.show_message("Erro", f"Erro ao salvar as configurações dos nomes: {str(e)}")
            raise

    @staticmethod
    def salvar_login(entries): # Função para salvar o login
        Usuario = entries['usuario_entry'].get()
        Senha = entries['senha_entry'].get()

        # Cria um dicionário com as configurações
        config = {
            "Usuario": Usuario, # Salva o nome do arquivo de estoque supper
            "Senha": Senha
        }

        # Salvar o caminho e os argumentos em um arquivo JSON
        try:
            with open('configuracoes_login.json', 'w') as f: # Abre o arquivo de configurações
                json.dump(config, f, indent=4)
            Utils.show_message("Sucesso", "o nome dos arquivos foram salvos.")
        except Exception as e:
            Utils.show_message("Erro", f"Erro ao salvar as configurações Login: {str(e)}")
            raise

    @staticmethod
    def salvar_telegram(entries): # Função para salvar o login
        Token = entries['Token_entry'].get()
        ChatID = entries['ChatID_entry'].get()
        Diretorio = entries['Diretorio_entry'].get()

        # Cria um dicionário com as configurações
        config = {
            "Token": Token, # Salva o nome do arquivo de estoque supper
            "ChatID": ChatID,
            "Diretorio": Diretorio
        }

        # Salvar o caminho e os argumentos em um arquivo JSON
        try:
            with open('configuracoes_telegram.json', 'w') as f: # Abre o arquivo de configurações
                json.dump(config, f, indent=4)
            Utils.show_message("Sucesso", "o nome dos arquivos foram salvos.")
        except Exception as e:
            Utils.show_message("Erro", f"Erro ao salvar as configurações Login: {str(e)}")
            raise

    @staticmethod
    def apicar_modelo_de_relatorio(caminho_imagem_lista): # Função para aplicar um modelo de relatório
        try:
            Utils.clicar_elemento('img/Modelo_Relatorio/SetaModelos.png', "Botão 'Seta Modelos'")
            time.sleep(10)  # Ajuste conforme necessário
            
            Utils.clicar_elemento('img/Modelo_Relatorio/SetaLista.png', "Botão 'Seta Lista Modelos'")
            time.sleep(10)  # Ajuste conforme necessário

            Utils.clicar_elemento(caminho_imagem_lista, "Modelo de Relatório, caminho imagem lista")
            time.sleep(10)  # Ajuste conforme necessário

        except Exception as e:
            logging.error(f"Ocorreu um erro ao aplicar o modelo de relatório: {e}")
            raise
    
    @staticmethod
    def salvar_arquivo(nome_arquivo): # Função para salvar um arquivo
        try:
            # Ler o caminho do arquivo de configuração
            with open('configuracoes_software.json', 'r') as f:
                config = json.load(f)
                caminho = config.get("caminho", "")
                if not caminho:
                    logging.error("O arquivo configuracoes_software.json não contém a chave 'caminho'.")
                    return

            # Simulação de operações de salvar arquivo usando pyautogui (adapte conforme necessário)
            # pyautogui.click('img/Salvar_Arquivo/ExportarExcel.png')
            Utils.clicar_elemento('img/Salvar_Arquivo/ExportarTXT.png', "Icone 'Exportar em TXT'")
            time.sleep(15)
            
            # pyautogui.click('img/Salvar_Arquivo/EsteComputador.png')
            Utils.clicar_elemento('img/Salvar_Arquivo/AreaDeTrabalho.png', "Icone 'Area de Trabalho'", 'img/Salvar_Arquivo/Objetos3D.png')
            time.sleep(5)
            
            # pyautogui.click('img/Salvar_Arquivo/EsteComputador_Pesquisa.png')
            Utils.clicar_elemento('img/Salvar_Arquivo/AreaDeTrabalho_Pesquisa.png', "Icone 'Area de Trabalho - barra de Pesquisa'", 'img/Salvar_Arquivo/Objetos3D_Pesquisa.png')
            time.sleep(5)
            pyautogui.write(caminho)
            pyautogui.press('enter')
            time.sleep(10)

            # pyautogui.click('img/Salvar_Arquivo/NomeDocumento.png')
            Utils.clicar_elemento('img/Salvar_Arquivo/NomeDocumento.png', "Campo 'Nome do Documento'")
            time.sleep(5)
            pyautogui.write(nome_arquivo)
            time.sleep(10)

            # pyautogui.click('img/Salvar_Arquivo/BotaoSalvar.png')
            Utils.clicar_elemento('img/Salvar_Arquivo/BotaoSalvar.png', "Botão 'Salvar'", 'img/Salvar_Arquivo/BotaoSalvar_v3.png')
            time.sleep(30)

            # pyautogui.click('img/Salvar_Arquivo/NaoAbrirArquivo.png')
            Utils.clicar_elemento('img/Salvar_Arquivo/NaoAbrirArquivo.png', "Botão 'Não/Ok'")
            time.sleep(10)

        except FileNotFoundError:
            Utils.show_message("Erro", "Arquivo de configurações não encontrado.")
        except Exception as e:
            logging.error(f"Ocorreu um erro ao salvar o arquivo: {e}")
            raise
    
    @staticmethod
    def salvar_arquivo_mrp(nome_arquivo): # Função para salvar um arquivo MRP
        try:
            # Ler o caminho do arquivo de configuração
            with open('configuracoes_software.json', 'r') as f:
                config = json.load(f)
                caminho = config.get("caminho_mrp", "")
                if not caminho:
                    logging.error("O arquivo configuracoes_software.json não contém a chave 'caminho_mrp'.")
                    return
                
            # Simulação de operações de salvar arquivo usando pyautogui (adapte conforme necessário)
            # pyautogui.click('img/Salvar_Arquivo/ExportarExcel.png')
            Utils.clicar_elemento('img/Salvar_Arquivo/ExportarExcel_v2.png', "Icone 'Exportar Excel'")
            time.sleep(10)
            
            # pyautogui.click('img/Salvar_Arquivo/EsteComputador.png')
            Utils.clicar_elemento('img/Salvar_Arquivo/AreaDeTrabalho.png', "Icone 'Este Computador'", 'img/Salvar_Arquivo/Objetos3D.png')
            time.sleep(5)
            
            # pyautogui.click('img/Salvar_Arquivo/EsteComputador_Pesquisa.png')
            Utils.clicar_elemento('img/Salvar_Arquivo/AreaDeTrabalho_Pesquisa.png', "Icone 'Este Computador - Barra de Pesquisa'", 'img/Salvar_Arquivo/Objetos3D_Pesquisa.png')
            time.sleep(2)
            pyautogui.write(caminho)
            pyautogui.press('enter')
            time.sleep(5)

            # pyautogui.click('img/Salvar_Arquivo/NomeDocumento.png')
            Utils.clicar_elemento('img/Salvar_Arquivo/NomeDocumento.png', "Campo 'Nome do Documento'")
            time.sleep(2)
            pyautogui.write(nome_arquivo)
            time.sleep(2)

            # pyautogui.click('img/Salvar_Arquivo/BotaoSalvar.png')
            Utils.clicar_elemento('img/Salvar_Arquivo/BotaoSalvar_v2.png', "Botão 'Salvar'", 'img/Salvar_Arquivo/BotaoSalvar_v3.png')
            time.sleep(15)

            # pyautogui.click('img/Salvar_Arquivo/NaoAbrirArquivo.png')
            Utils.clicar_elemento('img/Salvar_Arquivo/BotaoOK_v2.png', "Botão 'Não/Ok'")
            time.sleep(5)

        except FileNotFoundError:
            Utils.show_message("Erro", "Arquivo configurações_caminhos.txt não encontrado.")
        except Exception as e:
            logging.error(f"Ocorreu um erro ao salvar o arquivo: {e}")
            raise

    @staticmethod
    def aplicar_data_mes_atual(data_inicial_img, data_final_img): # Função para aplicar a data do mês atual
        try:
            data_atual = datetime.now()
            primeiro_dia_mes_atual = data_atual.replace(day=1).strftime('%d/%m/%Y')
            ultimo_dia_mes_atual = data_atual.replace(day=calendar.monthrange(data_atual.year, data_atual.month)[1]).strftime('%d/%m/%Y')

            Utils.clicar_elemento(data_inicial_img, "Campo 'Data Inicial'")
            time.sleep(2)  # Ajuste conforme necessário
            pyautogui.write(primeiro_dia_mes_atual)

            Utils.clicar_elemento(data_final_img, "Campo 'Data Final'")
            time.sleep(2)  # Ajuste conforme necessário
            pyautogui.write(ultimo_dia_mes_atual)

            Utils.clicar_elemento('img/Salvar_Arquivo/BotaoOk.png', "Botão 'Ok'")
            time.sleep(5)  # Ajuste conforme necessário

        except Exception as e:
            logging.error(f"Ocorreu um erro ao aplicar a data do mês atual: {e}")
            raise

    @staticmethod 
    def expand_All_Notes(coluna1, coluna2=None, coluna3=None): # Função para expandir todas as colunas do cubo
        try:
            Utils.clicar_elemento(coluna1, "Primeira Coluna")
            pyautogui.rightClick()
            time.sleep(5)  # Ajuste conforme necessário
            Utils.clicar_elemento('img/Expand_All_Notes/Expand_All_Notes.png', "Botão 'Expand All Notes'")
            time.sleep(5)  # Ajuste conforme necessário

            if coluna2:
                Utils.clicar_elemento(coluna2, "Segunda Coluna")
                pyautogui.rightClick()
                time.sleep(5)  # Ajuste conforme necessário
                Utils.clicar_elemento('img/Expand_All_Notes/Expand_All_Notes.png', "Botão 'Expand All Notes'")
                time.sleep(5)  # Ajuste conforme necessário

            if coluna3:
                Utils.clicar_elemento(coluna3, "Terceira Coluna")
                pyautogui.rightClick()
                time.sleep(5)  # Ajuste conforme necessário
                Utils.clicar_elemento('img/Expand_All_Notes/Expand_All_Notes.png', "Botão 'Expand All Notes'")
                time.sleep(5)  # Ajuste conforme necessário

        except Exception as e:
            logging.error(f"Ocorreu um erro ao expandir todas as colunas do cubo: {e}")
            raise

    @staticmethod
    def salvar_arquivo_parametros(nome_arquivo, exportar_excel_img): # Função para salvar um arquivo de parâmetros
        try:
            # Ler o caminho do arquivo de configuração
            with open('configuracoes_software.json', 'r') as f:
                config = json.load(f)
                caminho = config.get("caminho_parametros", "")
                if not caminho:
                    logging.error("O arquivo configuracoes_software.json não contém a chave 'caminho_parametros'.")
                    return
                
            now = datetime.now()
            sufixo_dia_mes_ano = now.strftime('_%d_%m_%Y')

            Utils.clicar_elemento(exportar_excel_img, "Icone 'Exportar Excel'")
            time.sleep(5)  # Ajuste conforme necessário
            
            Utils.clicar_elemento('img/Salvar_Arquivo/EsteComputador.png', "Icone 'Este Computador'", 'img/Salvar_Arquivo/Objetos3D.png')
            time.sleep(5)  # Ajuste conforme necessário
            
            Utils.clicar_elemento('img/Salvar_Arquivo/EsteComputador_Pesquisa.png', "Icone 'Este Computador - Barra de Pesquisa'", 'img/Salvar_Arquivo/Objetos3D_Pesquisa.png')
            time.sleep(2)  # Ajuste conforme necessário
            pyautogui.write(caminho+sufixo_dia_mes_ano)
            pyautogui.press('enter')
            time.sleep(5)  # Ajuste conforme necessário

            Utils.clicar_elemento('img/Salvar_Arquivo/NomeDocumento.png', "Campo 'Nome do Documento'")
            time.sleep(2)  # Ajuste conforme necessário

            pyautogui.write(nome_arquivo)
            time.sleep(2)  # Ajuste conforme necessário

            Utils.clicar_elemento('img/Salvar_Arquivo/SalvarParametros.png', "Botão 'Salvar'")
            time.sleep(5)  # Ajuste conforme necessário
        except Exception as e:
            logging.error(f"Ocorreu um erro ao salvar o arquivo: {e}")
            raise

    @staticmethod
    def salvar_arquivo_faturamento_peso(nome_arquivo): # Função para salvar um arquivo de faturamento peso
    
        try:
            # Ler o caminho do arquivo de configuração
            with open('configuracoes_software.json', 'r') as f:
                config = json.load(f)
                caminho = config.get("caminho_faturamento_peso", "")
                if not caminho:
                    logging.error("O arquivo configuracoes_software.json não contém a chave 'caminho_faturamento_peso'.")
                    return

            now = datetime.now()
            sufixo_ano_mes = now.strftime('_%Y_%m')

            # Simulação de operações de salvar arquivo usando pyautogui (adapte conforme necessário)
            # pyautogui.click('img/Salvar_Arquivo/ExportarExcel.png')
            Utils.clicar_elemento('img/Salvar_Arquivo/ExportarTXT.png', "Icone 'Exportar TXT'")
            time.sleep(5)
            
            # pyautogui.click('img/Salvar_Arquivo/EsteComputador.png')
            Utils.clicar_elemento('img/Salvar_Arquivo/AreaDeTrabalho.png', "Icone 'Este Computador'", 'img/Salvar_Arquivo/Objetos3D.png')
            time.sleep(5)
            
            # pyautogui.click('img/Salvar_Arquivo/EsteComputador_Pesquisa.png')
            Utils.clicar_elemento('img/Salvar_Arquivo/AreaDeTrabalho_Pesquisa.png', "Icone 'Este Computador Pesquisa'", 'img/Salvar_Arquivo/Objetos3D_Pesquisa.png')
            time.sleep(2)
            pyautogui.write(caminho)
            pyautogui.press('enter')
            time.sleep(5)

            # pyautogui.click('img/Salvar_Arquivo/NomeDocumento.png')
            Utils.clicar_elemento('img/Salvar_Arquivo/NomeDocumento.png', "Elemento 'Nome do Documento'")
            time.sleep(2)
            pyautogui.write(nome_arquivo+sufixo_ano_mes)
            time.sleep(2)

            # pyautogui.click('img/Salvar_Arquivo/BotaoSalvar.png')
            Utils.clicar_elemento('img/Salvar_Arquivo/BotaoSalvar.png', "Icone 'Botão Salvar'", 'img/Salvar_Arquivo/BotaoSalvar_v3.png')
            time.sleep(5)

            # pyautogui.click('img/Salvar_Arquivo/NaoAbrirArquivo.png')
            Utils.clicar_elemento('img/Salvar_Arquivo/NaoAbrirArquivo.png', "Botão 'Não/Ok'")
            time.sleep(5)

        except FileNotFoundError:
            Utils.show_message("Erro", "Arquivo de configurações não encontrado.")
        except Exception as e:
            logging.error(f"Ocorreu um erro ao salvar o arquivo: {e}")
            raise

    @staticmethod
    def salvar_arquivo_mov_saida_odf(nome_arquivo): # Função para salvar um arquivo de movimento de saída ODF
        try:
            # Ler o caminho do arquivo de configuração
            with open('configuracoes_software.json', 'r') as f:
                config = json.load(f)
                caminho = config.get("caminho_mov_odf", "")
                if not caminho:
                    logging.error("O arquivo configuracoes_software.json não contém a chave 'caminho_mov_odf'.")
                    return

            now = datetime.now()
            sufixo_ano_mes = now.strftime('_%Y_%m')

            # Simulação de operações de salvar arquivo usando pyautogui (adapte conforme necessário)
            # pyautogui.click('img/Salvar_Arquivo/ExportarExcel.png')
            Utils.clicar_elemento('img/Salvar_Arquivo/ExportarTXT.png', "Icone 'Exportar TXT'")
            time.sleep(5)
            
            # pyautogui.click('img/Salvar_Arquivo/EsteComputador.png')
            Utils.clicar_elemento('img/Salvar_Arquivo/AreaDeTrabalho.png', "Icone 'Este Computador'", 'img/Salvar_Arquivo/Objetos3D.png')
            time.sleep(5)
            
            # pyautogui.click('img/Salvar_Arquivo/EsteComputador_Pesquisa.png')
            Utils.clicar_elemento('img/Salvar_Arquivo/AreaDeTrabalho_Pesquisa.png', "Icone 'Este Computador - Barra de Pesquisa'", 'img/Salvar_Arquivo/Objetos3D_Pesquisa.png')
            time.sleep(2)
            pyautogui.write(caminho)
            pyautogui.press('enter')
            time.sleep(5)

            # pyautogui.click('img/Salvar_Arquivo/NomeDocumento.png')
            Utils.clicar_elemento('img/Salvar_Arquivo/NomeDocumento.png', "Campo 'Nome do Documento'")
            time.sleep(2)
            pyautogui.write(nome_arquivo+sufixo_ano_mes) # Adiciona o sufixo com o ano e mês atual
            time.sleep(2)

            # pyautogui.click('img/Salvar_Arquivo/BotaoSalvar.png')
            Utils.clicar_elemento('img/Salvar_Arquivo/BotaoSalvar.png', "Botão 'Salvar'", 'img/Salvar_Arquivo/BotaoSalvar_v3.png')
            time.sleep(5)

            # pyautogui.click('img/Salvar_Arquivo/NaoAbrirArquivo.png')
            Utils.clicar_elemento('img/Salvar_Arquivo/NaoAbrirArquivo.png', "Botão 'Não/Ok'")
            time.sleep(5)

        except FileNotFoundError:
            Utils.show_message("Erro", "Arquivo de configurações não encontrado.")
        except Exception as e:
            logging.error(f"Ocorreu um erro ao salvar o arquivo: {e}")
            raise

    @staticmethod
    def renomear_planilhas_arquivos_xls(caminho_pasta, novo_nome_planilha): # Função para renomear planilhas de arquivos .xls
        # Lista todos os arquivos na pasta fornecida
        try:
            for nome_arquivo in os.listdir(caminho_pasta):
                if nome_arquivo.endswith(".xls"):
                    caminho_arquivo = os.path.join(caminho_pasta, nome_arquivo)
                    
                    # Carrega o arquivo Excel
                    workbook = xlrd.open_workbook(caminho_arquivo, formatting_info=True)
                    novo_workbook = copy(workbook)
                    
                    # Renomeia todas as planilhas no arquivo Excel
                    for indice in range(len(workbook.sheet_names())):
                        novo_workbook.get_sheet(indice).set_name(novo_nome_planilha)
                    
                    # Salva as mudanças no arquivo Excel, sobrescrevendo o arquivo original
                    novo_workbook.save(caminho_arquivo)
                    logging.info(f"Planilhas renomeadas em: {nome_arquivo}")
        except Exception as e:
            logging.error(f"Ocorreu um erro ao renomear as planilhas dos arquivos {nome_arquivo}: {e}")
            raise

    @staticmethod
    def converter_xls_para_xlsx(caminho_pasta): # Função para converter arquivos .xls para .xlsx
        try:
        # Lista todos os arquivos na pasta fornecida
            for nome_arquivo in os.listdir(caminho_pasta):
                if nome_arquivo.endswith(".xls"):
                    caminho_arquivo = os.path.join(caminho_pasta, nome_arquivo)
                    
                    # Carrega o arquivo .xls
                    workbook_xls = xlrd.open_workbook(caminho_arquivo)
                    
                    # Cria um novo workbook .xlsx
                    workbook_xlsx = Workbook()
                    
                    for indice_planilha in range(workbook_xls.nsheets):
                        planilha_xls = workbook_xls.sheet_by_index(indice_planilha)
                        planilha_xlsx = workbook_xlsx.create_sheet(title=planilha_xls.name if indice_planilha != 0 else 'Sheet')
                        
                        for indice_linha in range(planilha_xls.nrows):
                            for indice_coluna in range(planilha_xls.ncols):
                                valor_celula = planilha_xls.cell_value(indice_linha, indice_coluna)
                                planilha_xlsx.cell(row=indice_linha + 1, column=indice_coluna + 1, value=valor_celula)
                    
                    # Remove a planilha padrão criada pelo openpyxl
                    if 'Sheet' in workbook_xlsx.sheetnames:
                        del workbook_xlsx['Sheet']
                    
                    # Salva o novo arquivo .xlsx
                    novo_caminho_arquivo = os.path.join(caminho_pasta, nome_arquivo.replace('.xls', '.xlsx'))
                    workbook_xlsx.save(novo_caminho_arquivo)
                    logging.info(f"Arquivo convertido: {novo_caminho_arquivo}")
        except Exception as e:
            logging.error(f"Ocorreu um erro ao converter os arquivos .xls para .xlsx: {e}")
            raise


class StdoutRedirector: # Classe para redirecionar a saída padrão para um widget de texto
    def __init__(self, text_widget):
        self.log_text = text_widget
        self.original_stdout = sys.stdout
        sys.stdout = self

    def __del__(self):
        sys.stdout = self.original_stdout

    def write(self, message):
        self.log_text.insert(ctk.END, message)
        self.log_text.see(ctk.END)

    def flush(self):
        pass


def configurar_logs(): # Função para configurar os logs
        # Cria a pasta de logs se não existir
        if not os.path.exists("logs"):
            os.makedirs("logs")
        
        # Configura um arquivo de log apenas para erros
        log_file = os.path.join("logs", f"log_erros_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.log")

        # Configura o logger apenas para capturar erros
        error_handler = logging.FileHandler(log_file)
        error_handler.setLevel(logging.INFO)  # Define que vai gravar apenas erros

        # Formato da mensagem de log
        formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
        error_handler.setFormatter(formatter)

        # Adiciona o handler de erro ao logger
        logger = logging.getLogger()
        logger.setLevel(logging.INFO)  # Define o nível global para INFO (vai aparecer no console)
        logger.addHandler(error_handler)

        # Exibe mensagens no console também
        console_handler = logging.StreamHandler(sys.stdout)
        console_handler.setLevel(logging.INFO)  # Exibe mensagens no console
        console_handler.setFormatter(formatter)
        logger.addHandler(console_handler)